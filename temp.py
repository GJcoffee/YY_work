code = '"null_$head,main_m_pk_accountingbook,main_m_pk_vouchertype,main_m_num,main_m_attachment,main_pk_prepared,main_m_prepareddate,m_explanation,m_accsubjcode,m_pk_currtype,m_debitamount,m_localdebitamount,m_groupdebitamount,m_globaldebitamount,unitname,m_price,m_debitquantity,m_creditquantity,m_creditamount,m_localcreditamount,m_groupcreditamount,m_globalcreditamount,m_checkno,m_checkdate,verifyno,verifydate,m_bankaccount,billtype,m_checkstyle,vat_pk_vatcountry,vat_pk_receivecountry,vat_businesscode,vat_pk_clientvatcode,vat_pk_suppliervatcode,vat_pk_taxcode,vat_direction,vat_moneyamount,m_excrate2,excrate3,excrate4,ass_1,ass_2,ass_3,ass_4,ass_5,ass_6,ass_7,ass_8,ass_9"'
other_code = '"cashflow,m_flag,cashflowcurr,m_money,m_moneymain,m_moneygroup,m_moneyglobal,cashflowName,cashflowCode"'
import openpyxl
from functools import reduce
from collections import defaultdict
from openpyxl import load_workbook


def read_data(read_path, key_row, row_index=3):
    ret_list, data = [], []
    excel = openpyxl.load_workbook(read_path, data_only=True)
    sheet = excel.active
    max_row = sheet.max_row
    key = [cell.value for cell in excel['Sheet1'][key_row]]
    for row in range(row_index, max_row + 1):
        if not sheet[f'A{row}'].value:
            return ret_list, row
        value = [cell.value for cell in excel['Sheet1'][row]]
        data = dict(zip(key, value))
        ret_list.append(data)
    return ret_list


def func(task, next_task):
    task[next_task['sum']].append(next_task)
    return task


def write_excel(model_file_path, tasks, others):
    vouchers = [tasks[term] for term in tasks]
    for voucher in vouchers:
        excel = openpyxl.load_workbook(model_file_path)
        sheet = excel['Sheet1']
        code_list = [x[code] for x in voucher]
        sheet.insert_rows(3, amount=len(voucher))
        row_index = 5 + len(voucher)
        # 写入前半部分凭证数据
        for iterm in voucher:
            cells = excel['Sheet1'][3 + voucher.index(iterm)]
            iterm.pop('sum')
            for cell, value in zip(cells, iterm):
                cell.value = iterm[value]
        # 写入后半部分凭证数据
        for iterm in code_list:
            if others.get(iterm):
                cells = excel['Sheet1'][row_index]
                cells = cells[:len(others[iterm])]
                row_index += 1
                for cell, value in zip(cells, others[iterm]):
                    cell.value = others[iterm][value]
        excel.save(r"""C:\Users\Datagrand\Desktop\model2.xlsx""")
        excel.close()


if __name__ == '__main__':
    tasks = []
    path = r"""C:\Users\Datagrand\Desktop\蜀电3月费用单据\3.6制单.xlsx"""
    model_path = r"""C:\Users\Datagrand\Desktop\model.xlsx"""
    task_list, row = read_data(path, 2)
    others = read_data(path, row + 1, row + 2)
    others = {term[other_code]: term for term in others}
    for term in task_list:
        term['sum'] = term['核算账簿'] + term['凭证类别编码'] + term['凭证号'] + term['制单日期']
        tasks.append(term)
    voucher_list = reduce(func, tasks, defaultdict(list))
    write_excel(model_path, voucher_list, others)
