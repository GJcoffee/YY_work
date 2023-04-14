import openpyxl
import xlrd
from functools import partial


def read_data(read_path):
    ret_list, data = [], []
    excel = openpyxl.load_workbook(read_path, data_only=True)
    sheet = excel.active
    max_row = sheet.max_row
    print(max_row)
    for row in range(1, max_row + 1):
        name = sheet[f'B{row}'].value
        num = sheet[f'A{row}'].value
        _dict = {name: num}
        ret_list.append(_dict)
    return ret_list


def read_data_xlrd(read_path):
    excel = xlrd.open_workbook(read_path)
    sheet = excel.sheets()[0]
    index_end = sheet.nrows
    ret_dict = {}
    for term in range(index_end):
        name = sheet.cell_value(term, 3)
        num = sheet.cell_value(term, 2)
        ret_dict[name] = num
    ret_dict ={iterm: ret_dict[iterm] for iterm in ret_dict if ret_dict[iterm]}
    print(ret_dict)
    return ret_dict


def select(exist, account):
    result = []
    for term in account:
        if term not in exist:
            result.append({term: account[term]})

    return result


def write_excel(password_error):
    path = r'C:\Users\Datagrand\Desktop\账号信息.xlsx'
    excel = openpyxl.load_workbook(path)
    sheet = excel.worksheets[0]
    for term in password_error:
        row = password_error.index(term)+1
        key = list(term.keys())[0]
        sheet[f'A{row}'].value = key
        sheet[f'B{row}'].value = term[key]
    excel.save(path)


def main():
    path = r'C:\Users\Datagrand\Desktop\国网nc账号.xlsx'
    read_path = r'C:\Users\Datagrand\Desktop\SGNC账号信息录入人.xls'
    exist = read_data(path)
    account = read_data_xlrd(read_path)
    result = select(exist, account)
    write_excel(result)

if __name__ == '__main__':
    main()


