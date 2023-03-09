import pypinyin
import openpyxl
import xlrd
import os


# 不带声调的(style=pypinyin.NORMAL)
def pinyin(word):
    s = ''
    for i in pypinyin.pinyin(word, style=pypinyin.NORMAL):
        s += ''.join(i)
    return s


def read_efficiency(read_path, index_end):
    """
    读取效率数据
    :param read_path: 文件路径
    :param index_end: 结束行
    :return: 姓名-效率
    """
    excel = xlrd.open_workbook(read_path)
    sheet = excel.sheets()[0]
    ret_dict = {}
    for term in range(index_end):
        name = sheet.cell_value(term, 0)
        num = sheet.cell_value(term, 1)
        ret_dict[pinyin(name).title()] = num
    print(ret_dict)
    return ret_dict


def read_data(read_path, name_index, index_begin, index_end):
    """读取请假数据"""
    excel = openpyxl.load_workbook(read_path, data_only=True)
    sheet = excel.active
    ret_dict = {}
    index_end += 1
    for term in range(index_begin, index_end):
        v, w, x, y = sheet[f'V{term}'].value, sheet[f'W{term}'].value, sheet[f'X{term}'].value, sheet[f'Y{term}'].value
        # print(v, w, x, y)
        name = pinyin(sheet[f'{name_index}{term}'].value)
        name = name.title() if '）' not in name else name.split('（')[0].title()
        ret_dict[name] = float(v if v else "0") + float(w if w else "0") + float(x if x else "0") + float(
            y if y else "0")
    print(ret_dict)
    return ret_dict


def insert_data(insert_path, holiday_dict, efficiency_dict):
    null_list = []
    name_list = []
    name_dict = {}
    wb = openpyxl.load_workbook(insert_path)
    sheet = wb['效率总表2021-2023']
    for i in range(1, sheet.max_row):
        # 获取表格中的全部姓名并去掉空格
        name = sheet[f'A{i}'].value
        name = name if not name else name.replace(' ', '').title()
        sheet[f'A{i}'].value = name
        name_list.append(name)
        name_dict[name] = i
    print(name_list)
    for term in holiday_dict:
        # 遍历请假数据，导入请假和效率数据
        if term in name_list:
            holiday = holiday_dict[term] / 8
            print(holiday)
            if not holiday:
                sheet[f'DA{name_dict[term]}'].value = None
                sheet[f'CZ{name_dict[term]}'].value = efficiency_dict.get(f'{term}')
                sheet[f'CY{name_dict[term]}'].value = '20'
                continue
            else:
                sheet[f'DA{name_dict[term]}'].value = 'Holiday=%2.3f' % holiday
                sheet[f'CZ{name_dict[term]}'].value = efficiency_dict.get(f'{term}')
                sheet[f'CY{name_dict[term]}'].value = '20'
                continue
        else:
            null_list.append({term: holiday_dict[term]})
    print('未导入数据：')
    for x in null_list:
        print(x)
    wb.save(rf'{os.path.expanduser("~")}\Desktop\Montly Efficiency_Color4Games2.xlsx')


if __name__ == '__main__':
    path = rf'{os.path.expanduser("~")}\Desktop'
    holiday_dict = read_data(os.path.join(path, '成都麟动科技有限公司_月度汇总_20230201-202302281.xlsx'), 'A', 5, 66)
    efficiency_dict = read_efficiency(os.path.join(path, '好了的.xls'), 51)
    insert_data(rf'{path}\Montly Efficiency_Color4Games.xlsx', holiday_dict, efficiency_dict)
